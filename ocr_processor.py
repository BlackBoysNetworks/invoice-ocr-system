#!/usr/bin/env python3
"""
OCR Processor para Facturas de Proveedores
Monitorea /srv/facturas/escaneadas, procesa PDFs/imágenes,
extrae datos y los registra en Excel.
"""

import os
import re
import shutil
import logging
import time
from datetime import datetime
from pathlib import Path

import pytesseract
from PIL import Image
import pdf2image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ── Configuración ────────────────────────────────────────────────────────────
SCAN_DIR      = Path("/srv/facturas/escaneadas")
PROCESSED_DIR = Path("/srv/facturas/procesadas")
EXCEL_PATH    = Path("/srv/facturas/facturas.xlsx")
LOG_PATH      = Path("/srv/facturas/ocr.log")

TESSERACT_LANG = "spa+eng"  # Español + Inglés
POPPLER_PATH   = None       # None = usa PATH del sistema

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ── Inicializar Excel ─────────────────────────────────────────────────────────
def init_excel():
    """Crea o abre el archivo Excel y asegura encabezados."""
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        return wb, ws

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = ["N° Factura", "Fecha", "Proveedor", "Monto Total", "Archivo", "Procesado"]
    header_fill  = PatternFill("solid", fgColor="1A237E")
    header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30
    col_widths = [18, 14, 30, 16, 40, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    return wb, ws


def append_to_excel(num_factura, fecha, proveedor, monto, archivo):
    """Agrega una fila al Excel con los datos de la factura."""
    wb, ws = init_excel()
    row = ws.max_row + 1

    even_fill = PatternFill("solid", fgColor="F5F5F5")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    fill = even_fill if row % 2 == 0 else odd_fill
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    values = [num_factura, fecha, proveedor, monto, archivo, datetime.now().strftime("%Y-%m-%d %H:%M")]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill   = fill
        cell.border = border
        cell.font   = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center" if col != 3 else "left",
                                   vertical="center")
        if col == 4:  # Monto
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")

    wb.save(EXCEL_PATH)
    log.info(f"Excel actualizado: Factura {num_factura} | {proveedor} | ${monto}")


# ── Extracción OCR ────────────────────────────────────────────────────────────
def image_to_text(image_path: Path) -> str:
    """Convierte imagen a texto con Tesseract."""
    img = Image.open(image_path)
    # Mejorar contraste para OCR
    img = img.convert("L")  # Escala de grises
    text = pytesseract.image_to_string(img, lang=TESSERACT_LANG,
                                       config="--psm 6")
    return text


def pdf_to_text(pdf_path: Path) -> str:
    """Convierte PDF a texto vía OCR."""
    try:
        pages = pdf2image.convert_from_path(str(pdf_path), dpi=300,
                                             poppler_path=POPPLER_PATH)
        texts = []
        for page in pages:
            page_bw = page.convert("L")
            t = pytesseract.image_to_string(page_bw, lang=TESSERACT_LANG,
                                             config="--psm 6")
            texts.append(t)
        return "\n".join(texts)
    except Exception as e:
        log.error(f"Error PDF→texto: {e}")
        return ""


def extract_invoice_number(text: str) -> str:
    """Extrae número de factura del texto OCR."""
    patterns = [
        r"FACTURA\s+ELECTR[OÓ]NICA[\s\S]{1,100}?N[°º2\s]*[:\-]?\s*(\d{2,12})",
        r"[Ff]actura\s*[Nn]°?\s*[:\-]?\s*([A-Z0-9]{1,5}-?\d{2,10})",
        r"[Ff]actura\s*[Nn]°?\s*[:\-]?\s*(\d{2,12})",
        r"N°?\s*[Ff]actura\s*[:\-]?\s*([A-Z0-9\-]{2,15})",
        r"[Ii]nvoice\s*[Nn]o\.?\s*[:\-]?\s*(\w{2,15})",
        r"FACTURA\s+N[°O]?\s*[:\-]?\s*([A-Z0-9\-]{2,15})",
        r"N[°º2]?\s*[:\-]?\s*(\d{2,10})\b",
    ]
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return m.group(1).strip().replace(" ", "")
    return "SIN-NUMERO"


def extract_date(text: str) -> str:
    """Extrae fecha de la factura."""
    patterns = [
        r"[Ff]echa\s*[:\-]?\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
        r"(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{4})",
        r"(\d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2})",
        r"[Dd]ate\s*[:\-]?\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
    ]
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return m.group(1).strip()
    return datetime.now().strftime("%d/%m/%Y")


def extract_supplier(text: str) -> str:
    """Extrae nombre del proveedor (Emisor o Señor(es))."""
    
    # Prioridad 1: Formato DTE Chile (Nombre del Emisor justo antes de su RUT)
    m = re.search(r"([\s\S]{5,100}?)\s+RUT\s*:\s*\d{1,2}\.?\d{3}\.?\d{3}-[\dkK]", text, re.IGNORECASE)
    if m:
        name = m.group(1).replace("\n", " ").strip()
        if len(name) > 3:
            return name[:60]

    # Prioridad 2: Buscar a quién se emitió (Señor(es)) o clásico Proveedor
    patterns = [
        r"Señor(?:\(es\))?\s*[:\-]?\s*(.+?)(?:\s+Fecha|\n)",
        r"Señores\s*[:\-]?\s*(.+?)(?:\s+Fecha|\n)",
        r"[Pp]roveedor\s*[:\-]?\s*(.+?)\n",
        r"[Ee]mpresa\s*[:\-]?\s*(.+?)\n",
        r"[Rr]azón\s+[Ss]ocial\s*[:\-]?\s*(.+?)\n",
        r"[Ss]upplier\s*[:\-]?\s*(.+?)\n",
        r"^([A-Z][A-ZÁÉÍÓÚ\s]{5,50}(?:S\.?A\.?|LTDA\.?|SpA|E\.?I\.?R\.?L\.?))",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.MULTILINE)
        if m:
            name = m.group(1).strip()
            if len(name) > 3:
                return name[:60]
    return "PROVEEDOR DESCONOCIDO"


def extract_total(text: str) -> float:
    """Extrae monto total de la factura."""
    patterns = [
        r"[Tt]otal\s+[Aa]\s+[Pp]agar\s*[:\$]?\s*([\d\.,]+)",
        r"[Tt]otal\s*[:\$]?\s*([\d\.,]+)",
        r"[Mm]onto\s+[Tt]otal\s*[:\$]?\s*([\d\.,]+)",
        r"TOTAL\s*\$?\s*([\d\.,]+)",
        r"[Ii]mporte\s+[Tt]otal\s*[:\$]?\s*([\d\.,]+)",
        r"[Tt]otal\s+[Ff]actura\s*[:\$]?\s*([\d\.,]+)",
        r"\$\s*([\d\.,]+)\s*$",
    ]
    candidates = []
    for pattern in patterns:
        for m in re.finditer(pattern, text, re.MULTILINE):
            raw = m.group(1).replace(".", "").replace(",", ".")
            try:
                val = float(re.sub(r"[^\d\.]", "", raw))
                if val > 0:
                    candidates.append(val)
            except ValueError:
                continue
    if candidates:
        return max(candidates)  # Generalmente el total es el mayor
    return 0.0


# ── Procesar archivo ──────────────────────────────────────────────────────────
def process_file(src_path: Path):
    """Procesa un archivo escaneado: OCR, renombra, actualiza Excel."""
    # Ignorar archivos ocultos de macOS
    if src_path.name.startswith("._") or src_path.name in [".DS_Store", "desktop.ini"]:
        log.debug(f"Ignorando archivo del sistema: {src_path.name}")
        return

    suffix = src_path.suffix.lower()
    if suffix not in [".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp"]:
        log.warning(f"Formato no soportado: {src_path.name}")
        return

    log.info(f"Procesando: {src_path.name}")

    try:
        # Extraer texto
        if suffix == ".pdf":
            text = pdf_to_text(src_path)
        else:
            text = image_to_text(src_path)

        if not text.strip():
            log.warning(f"No se pudo extraer texto de {src_path.name}")
            text = ""

        # Extraer datos
        num_factura = extract_invoice_number(text)
        fecha       = extract_date(text)
        proveedor   = extract_supplier(text)
        total       = extract_total(text)

        log.info(f"Datos extraídos → N°:{num_factura} | Fecha:{fecha} | "
                 f"Proveedor:{proveedor} | Total:${total:,.2f}")

        # Renombrar archivo
        new_name = f"Factura {num_factura}{suffix}"
        dest_path = PROCESSED_DIR / new_name

        # Evitar duplicados
        counter = 1
        while dest_path.exists():
            new_name = f"Factura {num_factura}_{counter}{suffix}"
            dest_path = PROCESSED_DIR / new_name
            counter += 1

        shutil.move(str(src_path), str(dest_path))
        log.info(f"Archivo movido a: {dest_path.name}")

        # Actualizar Excel
        append_to_excel(num_factura, fecha, proveedor, total, dest_path.name)

    except Exception as e:
        log.error(f"Error procesando {src_path.name}: {e}", exc_info=True)


# ── Watchdog Handler ──────────────────────────────────────────────────────────
def wait_for_file_ready(path: Path, stable_secs: int = 3, timeout: int = 120) -> bool:
    """
    Espera hasta que el archivo deje de crecer (copia de red completada).
    Comprueba el tamaño cada segundo; si permanece igual por `stable_secs`
    segundos consecutivos, considera que la copia terminó.
    Retorna False si supera timeout sin estabilizarse.
    """
    prev_size = -1
    stable_count = 0
    waited = 0
    while waited < timeout:
        try:
            curr_size = path.stat().st_size
        except FileNotFoundError:
            return False
        if curr_size == prev_size and curr_size > 0:
            stable_count += 1
            if stable_count >= stable_secs:
                return True
        else:
            stable_count = 0
        prev_size = curr_size
        time.sleep(1)
        waited += 1
    log.warning(f"Timeout esperando que termine la copia de {path.name}")
    return False


class InvoiceHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        path = Path(event.src_path)
        if wait_for_file_ready(path):
            process_file(path)

    def on_moved(self, event):
        if event.is_directory:
            return
        path = Path(event.dest_path)
        if wait_for_file_ready(path):
            process_file(path)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    SCAN_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    init_excel()

    log.info("=" * 60)
    log.info("Sistema OCR de Facturas iniciado")
    log.info(f"Monitoreando: {SCAN_DIR}")
    log.info(f"Procesadas en: {PROCESSED_DIR}")
    log.info(f"Excel: {EXCEL_PATH}")
    log.info("=" * 60)

    # Procesar archivos existentes al arrancar
    for f in SCAN_DIR.iterdir():
        if f.is_file():
            process_file(f)

    observer = Observer()
    observer.schedule(InvoiceHandler(), str(SCAN_DIR), recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(5)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
    log.info("Sistema OCR detenido.")


if __name__ == "__main__":
    main()
